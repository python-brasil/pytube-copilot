#!/usr/bin/env python3
"""
PyTube Copilot - Sistema de AI para analise personalizada de videos
Interface grafica com CustomTkinter seguindo prototipos de design
"""

import os
import sys
import uuid
import pickle
import threading
from pathlib import Path
from datetime import datetime
from tkinter import filedialog
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any
import customtkinter as ctk
from PIL import Image
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
import base64

# Libs para extracao de conteudo de arquivos
try:
    from PyPDF2 import PdfReader
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Configurar encoding UTF-8 para Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

# Importar funcoes do youtube_helper
from youtube_helper import (
    extract_audio,
    transcribe_audio,
    load_prompt,
    get_audio_duration,
    get_file_size_mb,
    format_duration,
    format_transcription_with_timestamps,
    SUPPORTED_VIDEO_FORMATS,
    SUPPORTED_AUDIO_FORMATS,
)

# ============================================================================
# CONFIGURACOES E CONSTANTES
# ============================================================================

APP_NAME = "PyTube Copilot"
APP_VERSION = "2.0.0"

# Diretorios
APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
PROMPTS_DIR = APP_DIR / "prompts"
OUTPUT_DIR = APP_DIR / "youtube_analysis"

# Arquivos de dados
DATABASE_FILE = DATA_DIR / "pytube_copilot.pkl"
CONFIG_FILE = DATA_DIR / "config.pkl"

# Criptografia
_INTERNAL_KEY = b"PyTube_Copilot_S3cur3_K3y_2025!@#"
_SALT = b"pytube_copilot_salt_v2"

# Cores do tema (Light Mode - Branco e Azul)
COLORS = {
    "bg_main": "#F5F7FA",           # Fundo principal cinza claro
    "bg_card": "#FFFFFF",           # Cards brancos
    "bg_header": "#1565C0",         # Header azul
    "primary": "#1976D2",           # Azul primario
    "primary_hover": "#1565C0",     # Azul hover
    "primary_light": "#BBDEFB",     # Azul claro
    "accent": "#2196F3",            # Azul accent
    "success": "#4CAF50",           # Verde sucesso
    "warning": "#FFC107",           # Amarelo warning
    "error": "#F44336",             # Vermelho erro
    "text_primary": "#212121",      # Texto escuro
    "text_secondary": "#757575",    # Texto secundario
    "text_on_primary": "#FFFFFF",   # Texto em fundos azuis
    "border": "#E0E0E0",            # Bordas cinza claro
    "border_dark": "#BDBDBD",       # Bordas mais escuras
    "button_yellow": "#FFC107",     # Botao amarelo
    "button_yellow_hover": "#FFB300",
    "button_red": "#EF5350",        # Botao vermelho
    "button_red_hover": "#E53935",
    "button_green": "#66BB6A",      # Botao verde
    "button_green_hover": "#4CAF50",
    "button_blue": "#42A5F5",       # Botao azul
    "button_blue_hover": "#1E88E5",
    "table_header": "#E3F2FD",      # Header de tabela azul claro
    "table_hover": "#F5F5F5",       # Hover em tabelas
}

# Modelos disponiveis
WHISPER_MODELS = ["whisper-1"]
GPT_MODELS = ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"]

# Formatos de documentos suportados
SUPPORTED_DOCUMENT_FORMATS = ['.txt', '.pdf', '.xlsx']

# Todos os formatos suportados (para exibicao)
ALL_SUPPORTED_FORMATS = SUPPORTED_VIDEO_FORMATS + SUPPORTED_AUDIO_FORMATS + SUPPORTED_DOCUMENT_FORMATS

# ============================================================================
# MODELOS DE DADOS (BANCO DE DADOS)
# ============================================================================

@dataclass
class Agent:
    """Modelo de dados para um agente"""
    id: str = field(default_factory=lambda: str(uuid.uuid4()))
    name: str = ""
    model: str = "gpt-4o-mini"
    prompt: str = ""
    is_system: bool = False  # Agentes do sistema nao podem ser deletados
    order: int = 0  # Ordem de processamento (menor = primeiro)
    created_at: datetime = field(default_factory=datetime.now)
    updated_at: datetime = field(default_factory=datetime.now)


@dataclass
class AgentResponse:
    """Resposta de um agente para um projeto"""
    agent_id: str = ""
    agent_name: str = ""
    content: str = ""
    prompt_used: str = ""  # Prompt processado com variaveis substituidas (para debug)
    created_at: datetime = field(default_factory=datetime.now)


@dataclass
class Project:
    """Modelo de dados para um projeto processado"""
    id: str = field(default_factory=lambda: str(uuid.uuid4()))
    file_name: str = ""
    file_path: str = ""
    file_size_mb: float = 0.0
    duration_seconds: float = 0.0
    format: str = ""
    status: str = "pending"  # pending, processing, completed, error
    error_message: str = ""
    transcription_text: str = ""
    transcription_segments: List[Dict] = field(default_factory=list)
    agent_responses: List[AgentResponse] = field(default_factory=list)
    created_at: datetime = field(default_factory=datetime.now)
    updated_at: datetime = field(default_factory=datetime.now)


@dataclass
class AppConfig:
    """Configuracoes da aplicacao"""
    api_key: str = ""
    whisper_model: str = "whisper-1"
    gpt_model: str = "gpt-4o-mini"
    language: str = "pt"
    last_directory: str = ""


@dataclass
class Pipeline:
    """Modelo de dados para um pipeline de processamento"""
    id: str = field(default_factory=lambda: str(uuid.uuid4()))
    name: str = ""
    agent_ids: List[str] = field(default_factory=list)
    is_system: bool = False
    created_at: datetime = field(default_factory=datetime.now)
    updated_at: datetime = field(default_factory=datetime.now)


@dataclass
class Database:
    """Banco de dados completo da aplicacao"""
    projects: List[Project] = field(default_factory=list)
    agents: List[Agent] = field(default_factory=list)
    pipelines: List['Pipeline'] = field(default_factory=list)
    config: AppConfig = field(default_factory=AppConfig)
    version: str = "2.0.0"


# ============================================================================
# GERENCIADOR DE BANCO DE DADOS
# ============================================================================

class DatabaseManager:
    """Gerencia persistencia de dados em arquivo .pkl com criptografia"""

    _fernet = None
    _database: Optional[Database] = None

    @classmethod
    def _get_fernet(cls) -> Fernet:
        """Obtem instancia do Fernet para criptografia"""
        if cls._fernet is None:
            kdf = PBKDF2HMAC(
                algorithm=hashes.SHA256(),
                length=32,
                salt=_SALT,
                iterations=100000,
            )
            key = base64.urlsafe_b64encode(kdf.derive(_INTERNAL_KEY))
            cls._fernet = Fernet(key)
        return cls._fernet

    @classmethod
    def _ensure_directories(cls):
        """Garante que os diretorios necessarios existam"""
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    @classmethod
    def _get_default_agents(cls) -> List[Agent]:
        """Retorna lista de agentes padrao do sistema"""
        agents = []

        # Agente Roteirista (ordem 1)
        roteirista_prompt = ""
        if (PROMPTS_DIR / "roteirista.txt").exists():
            roteirista_prompt = (PROMPTS_DIR / "roteirista.txt").read_text(encoding="utf-8")
        agents.append(Agent(
            id="agent_roteirista",
            name="ROTEIRISTA",
            model="gpt-4o-mini",
            prompt=roteirista_prompt,
            is_system=True,
            order=1
        ))

        # Agente Timestamps para descricao (ordem 2)
        timestamps_prompt = """Voce e um especialista em criar timestamps para descricao de videos do YouTube.

Sua tarefa e analisar a transcricao com timestamps e criar uma lista formatada de timestamps
para ser copiada diretamente na descricao do video.

## Formato de Saida:
0:00 - Introducao
[timestamp] - [titulo da secao]
...

Use os timestamps EXATOS da transcricao. Identifique mudancas de topico e crie titulos curtos e descritivos."""
        agents.append(Agent(
            id="agent_timestamps",
            name="TIMESTAMPS PARA DESCRICAO DO YOUTUBE",
            model="gpt-4o-mini",
            prompt=timestamps_prompt,
            is_system=True,
            order=2
        ))

        # Agente Resumo Geral (ordem 3)
        resumo_prompt = """Voce e um especialista em criar resumos concisos e informativos.

Sua tarefa e analisar a transcricao de um video e criar um resumo geral do conteudo.

## Instrucoes:
1. Identifique o tema principal do video
2. Liste os pontos mais importantes abordados
3. Crie um resumo em 3-5 paragrafos
4. Destaque informacoes uteis e insights

Seja claro, objetivo e mantenha a essencia do conteudo original."""
        agents.append(Agent(
            id="agent_resumo",
            name="RESUMO GERAL",
            model="gpt-4o-mini",
            prompt=resumo_prompt,
            is_system=True,
            order=3
        ))

        # Agente Titulos, Descricao e Tags (ordem 4) - usa resposta do roteirista
        seo_prompt = ""
        if (PROMPTS_DIR / "titulo_descricao_tags.txt").exists():
            seo_prompt = (PROMPTS_DIR / "titulo_descricao_tags.txt").read_text(encoding="utf-8")
        agents.append(Agent(
            id="agent_seo",
            name="TITULOS, DESCRICAO E TAGS",
            model="gpt-4o-mini",
            prompt=seo_prompt,
            is_system=True,
            order=4
        ))

        return agents

    @classmethod
    def _get_default_pipelines(cls) -> List[Pipeline]:
        """Retorna lista de pipelines padrao do sistema"""
        return [
            Pipeline(
                id="pipeline_youtube_helper",
                name="Youtube Helper",
                agent_ids=["agent_roteirista", "agent_timestamps", "agent_resumo", "agent_seo"],
                is_system=True
            )
        ]

    @classmethod
    def load(cls) -> Database:
        """Carrega banco de dados do arquivo"""
        cls._ensure_directories()

        if cls._database is not None:
            return cls._database

        if DATABASE_FILE.exists():
            try:
                encrypted_data = DATABASE_FILE.read_bytes()
                fernet = cls._get_fernet()
                decrypted_data = fernet.decrypt(encrypted_data)
                cls._database = pickle.loads(decrypted_data)

                # Garantir que agentes do sistema existam
                system_agent_ids = {a.id for a in cls._database.agents if a.is_system}
                default_agents = cls._get_default_agents()
                for agent in default_agents:
                    if agent.id not in system_agent_ids:
                        cls._database.agents.append(agent)

                # Backward compatibility: garantir que pipelines existam
                if not hasattr(cls._database, 'pipelines') or cls._database.pipelines is None:
                    cls._database.pipelines = cls._get_default_pipelines()
                else:
                    # Garantir que pipelines do sistema existam
                    system_pipeline_ids = {p.id for p in cls._database.pipelines if p.is_system}
                    default_pipelines = cls._get_default_pipelines()
                    for pipeline in default_pipelines:
                        if pipeline.id not in system_pipeline_ids:
                            cls._database.pipelines.append(pipeline)

                return cls._database
            except Exception as e:
                print(f"Erro ao carregar banco de dados: {e}")

        # Criar banco de dados padrao
        cls._database = Database(
            agents=cls._get_default_agents(),
            pipelines=cls._get_default_pipelines(),
            config=AppConfig(last_directory=str(Path.home()))
        )
        cls.save()
        return cls._database

    @classmethod
    def save(cls):
        """Salva banco de dados no arquivo"""
        if cls._database is None:
            return

        cls._ensure_directories()

        try:
            pickled_data = pickle.dumps(cls._database)
            fernet = cls._get_fernet()
            encrypted_data = fernet.encrypt(pickled_data)
            DATABASE_FILE.write_bytes(encrypted_data)
        except Exception as e:
            print(f"Erro ao salvar banco de dados: {e}")
            raise

    @classmethod
    def get_projects(cls, limit: int = None) -> List[Project]:
        """Retorna projetos ordenados por data (mais recente primeiro)"""
        db = cls.load()
        projects = sorted(db.projects, key=lambda p: p.created_at, reverse=True)
        if limit:
            return projects[:limit]
        return projects

    @classmethod
    def get_project(cls, project_id: str) -> Optional[Project]:
        """Retorna um projeto pelo ID"""
        db = cls.load()
        for project in db.projects:
            if project.id == project_id:
                return project
        return None

    @classmethod
    def add_project(cls, project: Project) -> Project:
        """Adiciona um novo projeto"""
        db = cls.load()
        db.projects.append(project)
        cls.save()
        return project

    @classmethod
    def update_project(cls, project: Project):
        """Atualiza um projeto existente"""
        db = cls.load()
        for i, p in enumerate(db.projects):
            if p.id == project.id:
                project.updated_at = datetime.now()
                db.projects[i] = project
                cls.save()
                return

    @classmethod
    def get_agents(cls, ordered: bool = True) -> List[Agent]:
        """Retorna todos os agentes, opcionalmente ordenados por order"""
        db = cls.load()
        # Garantir que todos os agentes tenham ordem definida
        for i, agent in enumerate(db.agents):
            if not hasattr(agent, 'order') or agent.order == 0:
                agent.order = i + 1
        if ordered:
            return sorted(db.agents, key=lambda a: a.order)
        return db.agents

    @classmethod
    def get_agent(cls, agent_id: str) -> Optional[Agent]:
        """Retorna um agente pelo ID"""
        db = cls.load()
        for agent in db.agents:
            if agent.id == agent_id:
                return agent
        return None

    @classmethod
    def add_agent(cls, agent: Agent) -> Agent:
        """Adiciona um novo agente"""
        db = cls.load()
        # Definir ordem como ultima posicao
        max_order = max((a.order for a in db.agents), default=0)
        agent.order = max_order + 1
        db.agents.append(agent)
        cls.save()
        return agent

    @classmethod
    def update_agent(cls, agent: Agent):
        """Atualiza um agente existente"""
        db = cls.load()
        for i, a in enumerate(db.agents):
            if a.id == agent.id:
                agent.updated_at = datetime.now()
                db.agents[i] = agent
                cls.save()
                return

    @classmethod
    def delete_agent(cls, agent_id: str) -> bool:
        """Deleta um agente (apenas agentes nao-sistema)"""
        db = cls.load()
        for i, agent in enumerate(db.agents):
            if agent.id == agent_id and not agent.is_system:
                db.agents.pop(i)
                cls._reorder_agents()
                cls.save()
                return True
        return False

    @classmethod
    def move_agent_up(cls, agent_id: str) -> bool:
        """Move um agente para cima na ordem de processamento"""
        db = cls.load()
        agents = sorted(db.agents, key=lambda a: a.order)

        for i, agent in enumerate(agents):
            if agent.id == agent_id and i > 0:
                # Trocar ordem com o anterior
                prev_agent = agents[i - 1]
                agent.order, prev_agent.order = prev_agent.order, agent.order
                cls.save()
                return True
        return False

    @classmethod
    def move_agent_down(cls, agent_id: str) -> bool:
        """Move um agente para baixo na ordem de processamento"""
        db = cls.load()
        agents = sorted(db.agents, key=lambda a: a.order)

        for i, agent in enumerate(agents):
            if agent.id == agent_id and i < len(agents) - 1:
                # Trocar ordem com o proximo
                next_agent = agents[i + 1]
                agent.order, next_agent.order = next_agent.order, agent.order
                cls.save()
                return True
        return False

    @classmethod
    def _reorder_agents(cls):
        """Reorganiza a ordem dos agentes sequencialmente"""
        db = cls.load()
        agents = sorted(db.agents, key=lambda a: a.order)
        for i, agent in enumerate(agents):
            agent.order = i + 1

    @classmethod
    def get_config(cls) -> AppConfig:
        """Retorna configuracoes"""
        db = cls.load()
        return db.config

    @classmethod
    def save_config(cls, config: AppConfig):
        """Salva configuracoes"""
        db = cls.load()
        db.config = config
        cls.save()

    @classmethod
    def has_api_key(cls) -> bool:
        """Verifica se existe API key configurada"""
        config = cls.get_config()
        return bool(config.api_key.strip())

    @classmethod
    def get_pipelines(cls) -> List[Pipeline]:
        """Retorna todos os pipelines"""
        db = cls.load()
        return list(db.pipelines)

    @classmethod
    def get_pipeline(cls, pipeline_id: str) -> Optional[Pipeline]:
        """Retorna um pipeline pelo ID"""
        db = cls.load()
        for pipeline in db.pipelines:
            if pipeline.id == pipeline_id:
                return pipeline
        return None

    @classmethod
    def add_pipeline(cls, pipeline: Pipeline) -> Pipeline:
        """Adiciona um novo pipeline"""
        db = cls.load()
        db.pipelines.append(pipeline)
        cls.save()
        return pipeline

    @classmethod
    def update_pipeline(cls, pipeline: Pipeline):
        """Atualiza um pipeline existente"""
        db = cls.load()
        for i, p in enumerate(db.pipelines):
            if p.id == pipeline.id:
                pipeline.updated_at = datetime.now()
                db.pipelines[i] = pipeline
                cls.save()
                return

    @classmethod
    def delete_pipeline(cls, pipeline_id: str) -> bool:
        """Deleta um pipeline (apenas pipelines nao-sistema)"""
        db = cls.load()
        for i, pipeline in enumerate(db.pipelines):
            if pipeline.id == pipeline_id and not pipeline.is_system:
                db.pipelines.pop(i)
                cls.save()
                return True
        return False


# ============================================================================
# FUNCOES DE EXTRACAO DE CONTEUDO DE ARQUIVOS
# ============================================================================

def extract_text_from_txt(file_path: Path) -> dict:
    """Extrai texto de arquivo .txt"""
    encodings = ['utf-8', 'latin-1', 'cp1252']
    text = ""
    for enc in encodings:
        try:
            text = file_path.read_text(encoding=enc)
            break
        except (UnicodeDecodeError, Exception):
            continue

    lines = text.split('\n')
    words = len(text.split())
    return {
        'text': text,
        'info': f"{len(lines)} linhas | {words} palavras",
        'type_label': 'TXT',
        'type': 'texto',
    }


def extract_text_from_pdf(file_path: Path) -> dict:
    """Extrai texto de arquivo .pdf"""
    if not HAS_PYPDF2:
        return {
            'text': '[ERRO: PyPDF2 nao instalado. Execute: pip install PyPDF2]',
            'info': 'Erro',
            'type_label': 'PDF',
            'type': 'documento',
        }

    reader = PdfReader(str(file_path))
    pages = len(reader.pages)
    text_parts = []
    for i, page in enumerate(reader.pages):
        page_text = page.extract_text() or ''
        if page_text.strip():
            text_parts.append(f"--- Pagina {i + 1} ---\n{page_text}")

    text = '\n\n'.join(text_parts)
    words = len(text.split())
    return {
        'text': text,
        'info': f"{pages} paginas | {words} palavras",
        'type_label': 'PDF',
        'type': 'documento',
    }


def extract_text_from_xlsx(file_path: Path) -> dict:
    """Extrai texto de arquivo .xlsx"""
    if not HAS_OPENPYXL:
        return {
            'text': '[ERRO: openpyxl nao instalado. Execute: pip install openpyxl]',
            'info': 'Erro',
            'type_label': 'XLSX',
            'type': 'planilha',
        }

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    sheets = wb.sheetnames
    text_parts = []
    total_rows = 0
    for sheet_name in sheets:
        ws = wb[sheet_name]
        text_parts.append(f"=== Planilha: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            line = ' | '.join(cells)
            if line.strip(' |'):
                text_parts.append(line)
                total_rows += 1
    wb.close()

    text = '\n'.join(text_parts)
    return {
        'text': text,
        'info': f"{len(sheets)} planilhas | {total_rows} linhas",
        'type_label': 'XLSX',
        'type': 'planilha',
    }


def extract_file_content(file_path: Path) -> dict:
    """Extrai conteudo de qualquer arquivo suportado"""
    ext = file_path.suffix.lower()
    if ext == '.txt':
        return extract_text_from_txt(file_path)
    elif ext == '.pdf':
        return extract_text_from_pdf(file_path)
    elif ext == '.xlsx':
        return extract_text_from_xlsx(file_path)
    else:
        return {
            'text': '',
            'info': '',
            'type_label': ext.upper().strip('.'),
            'type': 'desconhecido',
        }


def get_file_brief_info(file_path: Path) -> dict:
    """Retorna informacoes breves sobre um arquivo para exibicao na lista"""
    ext = file_path.suffix.lower()
    size_mb = get_file_size_mb(file_path)
    size_str = f"{size_mb:.1f} MB" if size_mb >= 1 else f"{size_mb * 1024:.0f} KB"

    if ext in SUPPORTED_VIDEO_FORMATS:
        return {
            'type': 'media', 'type_label': 'VIDEO', 'size': size_str,
            'extra': 'Calculando duracao...', 'color': '#1976D2',
            'needs_duration': True,
        }
    elif ext in SUPPORTED_AUDIO_FORMATS:
        return {
            'type': 'media', 'type_label': 'AUDIO', 'size': size_str,
            'extra': 'Calculando duracao...', 'color': '#4CAF50',
            'needs_duration': True,
        }
    elif ext == '.pdf':
        try:
            if HAS_PYPDF2:
                reader = PdfReader(str(file_path))
                pages = len(reader.pages)
                extra = f"{pages} paginas"
            else:
                extra = "PyPDF2 nao instalado"
        except Exception:
            extra = "Erro ao ler"
        return {
            'type': 'document', 'type_label': 'PDF', 'size': size_str,
            'extra': extra, 'color': '#F44336', 'needs_duration': False,
        }
    elif ext == '.txt':
        try:
            text = file_path.read_text(encoding='utf-8', errors='ignore')
            lines = len(text.split('\n'))
            words = len(text.split())
            extra = f"{lines} linhas | {words} palavras"
        except Exception:
            extra = ""
        return {
            'type': 'document', 'type_label': 'TXT', 'size': size_str,
            'extra': extra, 'color': '#757575', 'needs_duration': False,
        }
    elif ext == '.xlsx':
        try:
            if HAS_OPENPYXL:
                wb = openpyxl.load_workbook(str(file_path), read_only=True)
                sheets = len(wb.sheetnames)
                wb.close()
                extra = f"{sheets} planilhas"
            else:
                extra = "openpyxl nao instalado"
        except Exception:
            extra = "Erro ao ler"
        return {
            'type': 'document', 'type_label': 'XLSX', 'size': size_str,
            'extra': extra, 'color': '#217346', 'needs_duration': False,
        }
    else:
        return {
            'type': 'unknown', 'type_label': ext.upper().strip('.'), 'size': size_str,
            'extra': '', 'color': '#757575', 'needs_duration': False,
        }


# ============================================================================
# COMPONENTES DE UI
# ============================================================================

class HeaderFrame(ctk.CTkFrame):
    """Header da aplicacao - PyTube Copilot | Agents + Botoes"""

    def __init__(self, parent, on_history_click, on_settings_click, show_history_btn=True):
        super().__init__(parent, fg_color=COLORS["bg_header"], corner_radius=0, height=60)
        self.pack_propagate(False)

        # Container interno
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20)

        # Logo/Titulo a esquerda
        title_frame = ctk.CTkFrame(content, fg_color="transparent")
        title_frame.pack(side="left", fill="y")

        ctk.CTkLabel(
            title_frame,
            text="PyTube Copilot",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=COLORS["text_on_primary"]
        ).pack(side="left", pady=15)

        ctk.CTkLabel(
            title_frame,
            text="  |  Agents",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["primary_light"]
        ).pack(side="left", pady=15)

        # Botoes a direita
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(side="right", fill="y")

        # Botao Configuracoes (branco)
        ctk.CTkButton(
            btn_frame,
            text="Configuracoes",
            width=120,
            height=35,
            fg_color=COLORS["bg_card"],
            text_color=COLORS["bg_header"],
            hover_color=COLORS["border"],
            font=ctk.CTkFont(size=13),
            command=on_settings_click
        ).pack(side="right", pady=12, padx=(10, 0))

        # Botao Historico (amarelo)
        if show_history_btn:
            ctk.CTkButton(
                btn_frame,
                text="Historico",
                width=100,
                height=35,
                fg_color=COLORS["button_yellow"],
                text_color=COLORS["text_primary"],
                hover_color=COLORS["button_yellow_hover"],
                font=ctk.CTkFont(size=13),
                command=on_history_click
            ).pack(side="right", pady=12)


class RecentProjectsTable(ctk.CTkFrame):
    """Tabela de projetos recentes com scroll"""

    def __init__(self, parent, on_project_click, max_height: int = 250):
        super().__init__(parent, fg_color="transparent")
        self.on_project_click = on_project_click

        # Titulo
        ctk.CTkLabel(
            self,
            text="Recentes",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS["text_primary"],
            anchor="w"
        ).pack(fill="x")

        # Container da tabela com altura maxima
        self.table_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=8, border_width=1, border_color=COLORS["border"])
        self.table_frame.pack(fill="x")

        # Header da tabela
        self._create_table_header()

        # Rows container com scroll
        self.rows_frame = ctk.CTkScrollableFrame(
            self.table_frame,
            fg_color="transparent",
            height=max_height,
            scrollbar_button_color=COLORS["border_dark"],
            scrollbar_button_hover_color=COLORS["primary"]
        )
        self.rows_frame.pack(fill="x", padx=2, pady=2)

        self.refresh()

    def _create_table_header(self):
        """Cria header da tabela"""
        header = ctk.CTkFrame(self.table_frame, fg_color=COLORS["table_header"], height=35)
        header.pack(fill="x", padx=2, pady=(2, 0))
        header.pack_propagate(False)

        columns = [("Titulo", 0.4), ("Tamanho", 0.15), ("Duracao", 0.15), ("Data", 0.3)]

        for col_name, weight in columns:
            ctk.CTkLabel(
                header,
                text=col_name,
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=COLORS["primary"],
                anchor="w"
            ).pack(side="left", fill="both", expand=True, padx=10, pady=5)

    def refresh(self):
        """Atualiza lista de projetos"""
        # Limpar rows existentes
        for widget in self.rows_frame.winfo_children():
            widget.destroy()

        # Carregar projetos recentes (sem limite, scroll permite ver todos)
        projects = DatabaseManager.get_projects(limit=20)

        if not projects:
            ctk.CTkLabel(
                self.rows_frame,
                text="Nenhum projeto encontrado",
                font=ctk.CTkFont(size=13),
                text_color=COLORS["text_secondary"]
            ).pack(pady=30)
            return

        for project in projects:
            self._create_row(project)

    def _create_row(self, project: Project):
        """Cria uma linha da tabela"""
        row = ctk.CTkFrame(self.rows_frame, fg_color="transparent", height=40, cursor="hand2")
        row.pack(fill="x", pady=1)
        row.pack_propagate(False)

        # Bind para hover e click
        row.bind("<Enter>", lambda e, r=row: r.configure(fg_color=COLORS["table_hover"]))
        row.bind("<Leave>", lambda e, r=row: r.configure(fg_color="transparent"))
        row.bind("<Button-1>", lambda e, p=project: self.on_project_click(p))

        # Dados
        duration_str = format_duration(project.duration_seconds) if project.duration_seconds else "--:--"
        date_str = project.created_at.strftime("%d/%m/%Y")

        data = [
            (project.file_name[:40] + "..." if len(project.file_name) > 40 else project.file_name, 0.4),
            (f"{project.file_size_mb:.0f} MB", 0.15),
            (duration_str, 0.15),
            (date_str, 0.3)
        ]

        for text, weight in data:
            label = ctk.CTkLabel(
                row,
                text=text,
                font=ctk.CTkFont(size=12),
                text_color=COLORS["text_primary"],
                anchor="w"
            )
            label.pack(side="left", fill="both", expand=True, padx=10, pady=8)
            label.bind("<Button-1>", lambda e, p=project: self.on_project_click(p))


class VideoInfoCard(ctk.CTkFrame):
    """Card com informacoes do video selecionado"""

    def __init__(self, parent, on_remove, on_process):
        super().__init__(parent, fg_color=COLORS["bg_card"], corner_radius=12, border_width=1, border_color=COLORS["border"])
        self.on_remove = on_remove
        self.on_process = on_process

        # Container horizontal
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20, pady=20)

        # Thumbnail placeholder
        self.thumb_frame = ctk.CTkFrame(content, fg_color=COLORS["primary_light"], width=160, height=100, corner_radius=8)
        self.thumb_frame.pack(side="left", padx=(0, 20))
        self.thumb_frame.pack_propagate(False)

        ctk.CTkLabel(
            self.thumb_frame,
            text="Thumbnail",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["primary"]
        ).pack(expand=True)

        # Info do arquivo
        info_frame = ctk.CTkFrame(content, fg_color="transparent")
        info_frame.pack(side="left", fill="both", expand=True)

        self.title_label = ctk.CTkLabel(
            info_frame,
            text="",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLORS["text_primary"],
            anchor="w"
        )
        self.title_label.pack(anchor="w")

        self.size_label = ctk.CTkLabel(
            info_frame,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=COLORS["text_secondary"],
            anchor="w"
        )
        self.size_label.pack(anchor="w", pady=(5, 0))

        self.duration_label = ctk.CTkLabel(
            info_frame,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=COLORS["text_secondary"],
            anchor="w"
        )
        self.duration_label.pack(anchor="w", pady=(2, 0))

        self.format_label = ctk.CTkLabel(
            info_frame,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=COLORS["text_secondary"],
            anchor="w"
        )
        self.format_label.pack(anchor="w", pady=(2, 0))

        # Botoes
        btn_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
        btn_frame.pack(anchor="w", pady=(15, 0))

        self.remove_btn = ctk.CTkButton(
            btn_frame,
            text="Remover",
            width=100,
            height=35,
            fg_color=COLORS["button_red"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["button_red_hover"],
            font=ctk.CTkFont(size=13),
            command=on_remove
        )
        self.remove_btn.pack(side="left", padx=(0, 10))

        self.process_btn = ctk.CTkButton(
            btn_frame,
            text="Processar",
            width=100,
            height=35,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=13),
            command=on_process
        )
        self.process_btn.pack(side="left")

    def set_file_info(self, file_path: Path, size_mb: float, duration: float):
        """Define informacoes do arquivo"""
        is_video = file_path.suffix.lower() in SUPPORTED_VIDEO_FORMATS
        file_type = "video" if is_video else "audio"

        self.title_label.configure(text=file_path.name)
        self.size_label.configure(text=f"Tamanho: {size_mb:.0f} MB")
        self.duration_label.configure(text=f"Duracao: {format_duration(duration)}")
        self.format_label.configure(text=f"Formato: {file_path.suffix.lower()} ({file_type})")

    def set_buttons_enabled(self, enabled: bool):
        """Habilita/desabilita botoes"""
        state = "normal" if enabled else "disabled"
        self.remove_btn.configure(state=state)
        self.process_btn.configure(state=state)


class ProcessingProgress(ctk.CTkFrame):
    """Barra de progresso de processamento"""

    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")

        self.pipeline_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["primary"]
        )
        self.pipeline_label.pack(pady=(15, 5))

        self.progress_bar = ctk.CTkProgressBar(
            self,
            width=500,
            height=20,
            progress_color=COLORS["primary"],
            fg_color=COLORS["border"]
        )
        self.progress_bar.pack(pady=(5, 10))
        self.progress_bar.set(0)

        self.status_label = ctk.CTkLabel(
            self,
            text="0% - Iniciando...",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["text_secondary"]
        )
        self.status_label.pack()

    def set_pipeline_name(self, name: str):
        """Define o nome do pipeline em execucao"""
        self.pipeline_label.configure(text=f"Pipeline: {name}")

    def update_progress(self, progress: float, message: str):
        """Atualiza progresso"""
        self.progress_bar.set(progress)
        self.status_label.configure(text=f"{int(progress * 100)}% - {message}")


class ProcessingResult(ctk.CTkFrame):
    """Resultado do processamento (sucesso ou erro)"""

    def __init__(self, parent, on_access):
        super().__init__(parent, fg_color="transparent")
        self.on_access = on_access

        self.message_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["success"]
        )
        self.message_label.pack(pady=(20, 15))

        self.access_btn = ctk.CTkButton(
            self,
            text="Acessar",
            width=120,
            height=40,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=14),
            command=on_access
        )
        self.access_btn.pack()

    def show_success(self, message: str = "SUCESSO AO PROCESSAR CLIQUE ABAIXO PARA VER O RESULTADO"):
        """Mostra mensagem de sucesso"""
        self.message_label.configure(text=message, text_color=COLORS["success"])
        self.access_btn.pack()

    def show_error(self, message: str):
        """Mostra mensagem de erro"""
        self.message_label.configure(text=f"ERRO AO PROCESSAR: {message}", text_color=COLORS["error"])
        self.access_btn.pack_forget()


class FileListPanel(ctk.CTkFrame):
    """Painel com lista de arquivos selecionados, botoes Remover e botao Processar"""

    def __init__(self, parent, on_remove_callback, on_process_callback):
        super().__init__(parent, fg_color="transparent")
        self.on_remove_callback = on_remove_callback
        self.on_process_callback = on_process_callback
        self.file_entries = []  # List of dicts: {path, info, frame}

        # Container da lista de arquivos
        self.list_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.list_frame.pack(fill="x")

        # Frame inferior: Pipeline selector + Botao Processar
        self.bottom_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.bottom_frame.pack(fill="x", pady=(10, 0))

        # Label Pipeline
        ctk.CTkLabel(
            self.bottom_frame,
            text="Pipeline:",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left", padx=(0, 5))

        # ComboBox de Pipeline
        self.pipeline_combo = ctk.CTkComboBox(
            self.bottom_frame,
            width=220,
            height=40,
            border_color=COLORS["border"],
            fg_color=COLORS["bg_card"],
            button_color=COLORS["primary"],
            button_hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=13),
            state="readonly"
        )
        self.pipeline_combo.pack(side="left", padx=(0, 10))
        self._pipeline_map = {}  # {name: pipeline}
        self._refresh_pipelines()

        # Botao Processar
        self.process_btn = ctk.CTkButton(
            self.bottom_frame,
            text="Processar",
            height=40,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=15, weight="bold"),
            command=on_process_callback
        )
        self.process_btn.pack(side="left", fill="x", expand=True)

    def add_file(self, file_path: Path, file_info: dict):
        """Adiciona um arquivo a lista"""
        # Verificar se ja existe
        for entry in self.file_entries:
            if entry['path'] == file_path:
                return

        # Criar card do arquivo
        card = ctk.CTkFrame(
            self.list_frame,
            fg_color=COLORS["bg_card"],
            corner_radius=8,
            border_width=1,
            border_color=COLORS["border"],
            height=60
        )
        card.pack(fill="x", pady=(0, 5))
        card.pack_propagate(False)

        content = ctk.CTkFrame(card, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=12, pady=8)

        # Badge de tipo (colorido)
        type_badge = ctk.CTkLabel(
            content,
            text=f" {file_info['type_label']} ",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#FFFFFF",
            fg_color=file_info.get('color', COLORS["primary"]),
            corner_radius=4,
            width=55
        )
        type_badge.pack(side="left", padx=(0, 10))

        # Info do arquivo (nome + detalhes)
        info_frame = ctk.CTkFrame(content, fg_color="transparent")
        info_frame.pack(side="left", fill="both", expand=True)

        name_text = file_path.name
        if len(name_text) > 45:
            name_text = name_text[:42] + "..."

        ctk.CTkLabel(
            info_frame,
            text=name_text,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"],
            anchor="w"
        ).pack(anchor="w")

        detail_text = file_info['size']
        if file_info.get('extra'):
            detail_text += f"  |  {file_info['extra']}"

        detail_label = ctk.CTkLabel(
            info_frame,
            text=detail_text,
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"],
            anchor="w"
        )
        detail_label.pack(anchor="w")

        # Botao Remover
        remove_btn = ctk.CTkButton(
            content,
            text="Remover",
            width=80,
            height=30,
            fg_color=COLORS["button_red"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["button_red_hover"],
            font=ctk.CTkFont(size=12),
            command=lambda fp=file_path: self._remove_file(fp)
        )
        remove_btn.pack(side="right", padx=(10, 0))

        entry = {
            'path': file_path,
            'info': file_info,
            'frame': card,
            'detail_label': detail_label,
            'remove_btn': remove_btn,
        }
        self.file_entries.append(entry)

    def _remove_file(self, file_path: Path):
        """Remove um arquivo da lista"""
        for i, entry in enumerate(self.file_entries):
            if entry['path'] == file_path:
                entry['frame'].destroy()
                self.file_entries.pop(i)
                break
        self.on_remove_callback(file_path)

    def update_file_extra(self, file_path: Path, extra_text: str):
        """Atualiza o texto extra de um arquivo (ex: duracao calculada)"""
        for entry in self.file_entries:
            if entry['path'] == file_path:
                current = entry['info']['size']
                if extra_text:
                    current += f"  |  {extra_text}"
                entry['detail_label'].configure(text=current)
                entry['info']['extra'] = extra_text
                break

    def get_files(self) -> List[Path]:
        """Retorna lista de arquivos selecionados"""
        return [e['path'] for e in self.file_entries]

    def has_files(self) -> bool:
        """Verifica se ha arquivos na lista"""
        return len(self.file_entries) > 0

    def clear(self):
        """Remove todos os arquivos"""
        for entry in self.file_entries:
            entry['frame'].destroy()
        self.file_entries.clear()

    def _refresh_pipelines(self):
        """Atualiza lista de pipelines no combo"""
        pipelines = DatabaseManager.get_pipelines()
        self._pipeline_map = {p.name: p for p in pipelines}
        names = [p.name for p in pipelines]
        self.pipeline_combo.configure(values=names)
        if names:
            self.pipeline_combo.set(names[0])

    def get_selected_pipeline(self) -> Optional['Pipeline']:
        """Retorna o pipeline selecionado"""
        selected_name = self.pipeline_combo.get()
        return self._pipeline_map.get(selected_name)

    def refresh_pipeline_list(self):
        """Metodo publico para forcar atualizacao dos pipelines"""
        self._refresh_pipelines()

    def set_enabled(self, enabled: bool):
        """Habilita/desabilita todos os botoes"""
        state = "normal" if enabled else "disabled"
        self.process_btn.configure(state=state)
        self.pipeline_combo.configure(state="readonly" if enabled else "disabled")
        for entry in self.file_entries:
            entry['remove_btn'].configure(state=state)


# ============================================================================
# TELAS PRINCIPAIS
# ============================================================================

class HomeScreen(ctk.CTkFrame):
    """Tela inicial - Tela #01 e #02"""

    def __init__(self, parent, app):
        super().__init__(parent, fg_color=COLORS["bg_main"])
        self.app = app
        self.selected_files = []  # Lista de Path
        self.file_durations = {}  # {Path: float}
        self.current_projects = []  # Lista de Projects sendo processados
        self.last_processed_project = None
        self.is_processing = False
        self.current_pipeline = None  # Pipeline selecionado para processamento

        self._create_widgets()

    def _create_widgets(self):
        """Cria widgets da tela"""
        # Header
        self.header = HeaderFrame(
            self,
            on_history_click=self.app.show_history,
            on_settings_click=self.app.show_settings
        )
        self.header.pack(fill="x")

        # Container principal com scroll para tela inteira
        self.main_container = ctk.CTkScrollableFrame(
            self,
            fg_color="transparent",
            scrollbar_button_color=COLORS["border_dark"],
            scrollbar_button_hover_color=COLORS["primary"]
        )
        self.main_container.pack(fill="both", expand=True, padx=30, pady=(10, 10))

        # Titulo
        ctk.CTkLabel(
            self.main_container,
            text="Selecione arquivos para comecar",
            font=ctk.CTkFont(size=20),
            text_color=COLORS["text_primary"]
        ).pack(pady=(0, 5))

        # Botao Selecionar Arquivo
        self.select_btn = ctk.CTkButton(
            self.main_container,
            text="Selecionar arquivo",
            width=200,
            height=45,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=15),
            command=self._browse_file
        )
        self.select_btn.pack()

        # Texto de formatos suportados
        formats_text = "Formatos suportados: MP4, MKV, AVI, MOV, MP3, WAV, TXT, PDF, XLSX"
        ctk.CTkLabel(
            self.main_container,
            text=formats_text,
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        ).pack(pady=(4, 0))

        # Painel de lista de arquivos (oculto inicialmente)
        self.file_list_panel = FileListPanel(
            self.main_container,
            on_remove_callback=self._on_file_removed,
            on_process_callback=self._start_processing
        )

        # Container para progresso/resultado
        self.progress_container = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.progress_container.pack(fill="x")

        self.progress_widget = ProcessingProgress(self.progress_container)
        self.result_widget = ProcessingResult(self.progress_container, on_access=self._access_result)

        # Tabela de recentes (com altura fixa e scroll interno)
        self.recents_table = RecentProjectsTable(
            self.main_container,
            on_project_click=self._on_project_click,
            max_height=280
        )
        self.recents_table.pack(fill="x", pady=(8, 0))

    def _browse_file(self):
        """Abre dialogo para selecionar arquivo(s)"""
        all_formats = SUPPORTED_VIDEO_FORMATS + SUPPORTED_AUDIO_FORMATS + SUPPORTED_DOCUMENT_FORMATS
        filetypes = [
            ("Todos suportados", " ".join(f"*{ext}" for ext in all_formats)),
            ("Videos", " ".join(f"*{ext}" for ext in SUPPORTED_VIDEO_FORMATS)),
            ("Audios", " ".join(f"*{ext}" for ext in SUPPORTED_AUDIO_FORMATS)),
            ("Documentos", " ".join(f"*{ext}" for ext in SUPPORTED_DOCUMENT_FORMATS)),
        ]

        config = DatabaseManager.get_config()
        initial_dir = config.last_directory or str(Path.home())

        file_paths = filedialog.askopenfilenames(
            title="Selecionar Arquivo(s)",
            initialdir=initial_dir,
            filetypes=filetypes
        )

        if file_paths:
            # Salvar diretorio do primeiro arquivo
            first = Path(file_paths[0])
            config.last_directory = str(first.parent)
            DatabaseManager.save_config(config)

            for fp in file_paths:
                path = Path(fp)
                if path not in self.selected_files:
                    self.selected_files.append(path)
                    self._add_file_to_list(path)

            # Mostrar painel de arquivos se tem arquivos
            if self.selected_files:
                self.file_list_panel.pack(fill="x", pady=(8, 0), before=self.progress_container)

                # Ocultar widgets de progresso/resultado
                self.progress_widget.pack_forget()
                self.result_widget.pack_forget()

    def _add_file_to_list(self, file_path: Path):
        """Adiciona arquivo ao painel e carrega info"""
        file_info = get_file_brief_info(file_path)
        self.file_list_panel.add_file(file_path, file_info)

        # Se precisa calcular duracao (media), faz em thread
        if file_info.get('needs_duration'):
            threading.Thread(
                target=self._load_file_duration,
                args=(file_path,),
                daemon=True
            ).start()

    def _load_file_duration(self, file_path: Path):
        """Carrega duracao de arquivo de midia em background"""
        try:
            duration = get_audio_duration(file_path)
            self.file_durations[file_path] = duration
            duration_str = format_duration(duration)
            self.after(0, lambda: self.file_list_panel.update_file_extra(
                file_path, f"Duracao: {duration_str}"
            ))
        except Exception:
            self.file_durations[file_path] = 0
            self.after(0, lambda: self.file_list_panel.update_file_extra(
                file_path, "Duracao: Erro"
            ))

    def _on_file_removed(self, file_path: Path):
        """Callback quando arquivo e removido da lista"""
        if file_path in self.selected_files:
            self.selected_files.remove(file_path)
        if file_path in self.file_durations:
            del self.file_durations[file_path]

        # Ocultar painel se nao tem mais arquivos
        if not self.selected_files:
            self.file_list_panel.pack_forget()

    def _start_processing(self):
        """Inicia processamento de todos os arquivos"""
        if not self.selected_files or self.is_processing:
            return

        # Verificar API key (necessaria para media files que usam transcricao)
        has_media = any(
            f.suffix.lower() in SUPPORTED_VIDEO_FORMATS + SUPPORTED_AUDIO_FORMATS
            for f in self.selected_files
        )
        if has_media and not DatabaseManager.has_api_key():
            self.app.show_settings(show_warning=True)
            return

        self.is_processing = True

        # Obter pipeline selecionado
        self.current_pipeline = self.file_list_panel.get_selected_pipeline()
        if not self.current_pipeline:
            return

        # Desabilitar botoes
        self.file_list_panel.set_enabled(False)
        self.select_btn.configure(state="disabled")

        # Ocultar tabela de recentes
        self.recents_table.pack_forget()

        # Mostrar progresso com nome do pipeline
        self.progress_widget.set_pipeline_name(self.current_pipeline.name)
        self.progress_widget.pack()
        self.result_widget.pack_forget()

        # Criar projetos para cada arquivo
        self.current_projects = []
        for file_path in self.selected_files:
            project = Project(
                file_name=file_path.name,
                file_path=str(file_path),
                file_size_mb=get_file_size_mb(file_path),
                duration_seconds=self.file_durations.get(file_path, 0),
                format=file_path.suffix.lower(),
                status="processing"
            )
            DatabaseManager.add_project(project)
            self.current_projects.append(project)

        # Iniciar processamento em thread
        threading.Thread(target=self._process_all_files, daemon=True).start()

    def _process_all_files(self):
        """Processa todos os arquivos em thread separada"""
        total_files = len(self.current_projects)
        has_error = False

        for file_idx, project in enumerate(self.current_projects):
            file_path = Path(project.file_path)
            file_prefix = f"[{file_idx + 1}/{total_files}] {file_path.name[:30]}"

            # Progresso base para este arquivo
            base_progress = file_idx / total_files
            file_progress_range = 1.0 / total_files

            try:
                self._process_single_file(
                    project, file_path, file_prefix,
                    base_progress, file_progress_range
                )
                project.status = "completed"
                DatabaseManager.update_project(project)
                self.last_processed_project = project

            except Exception as e:
                error_msg = str(e)
                project.status = "error"
                project.error_message = error_msg
                DatabaseManager.update_project(project)
                has_error = True

        # Finalizar
        self._update_progress(1.0, "Concluido!")

        if has_error and not self.last_processed_project:
            self.after(0, lambda: self._show_error("Erro ao processar um ou mais arquivos"))
        else:
            self.after(500, self._show_success)

        self.is_processing = False
        self.after(0, lambda: self.select_btn.configure(state="normal"))

    def _process_single_file(self, project: Project, file_path: Path,
                              prefix: str, base_progress: float, progress_range: float):
        """Processa um unico arquivo"""
        from openai import OpenAI

        ext = file_path.suffix.lower()
        is_video = ext in SUPPORTED_VIDEO_FORMATS
        is_audio = ext in SUPPORTED_AUDIO_FORMATS
        is_media = is_video or is_audio
        is_document = ext in SUPPORTED_DOCUMENT_FORMATS

        config = DatabaseManager.get_config()

        if is_media:
            # FLUXO MEDIA: Extrair audio -> Transcrever -> Agentes
            client = OpenAI(api_key=config.api_key)

            # Etapa 1: Extrair audio (se video)
            self._update_progress(base_progress + progress_range * 0.05,
                                  f"{prefix} - Extraindo audio...")

            if is_video:
                audio_path = extract_audio(file_path)
            else:
                audio_path = file_path

            # Obter duracao
            duration = get_audio_duration(audio_path)
            project.duration_seconds = duration

            # Etapa 2: Transcricao
            self._update_progress(base_progress + progress_range * 0.15,
                                  f"{prefix} - Transcrevendo...")
            transcription_data = transcribe_audio(client, audio_path, config.language)

            project.transcription_text = transcription_data['text']
            project.transcription_segments = transcription_data['segments']

            # Adicionar transcricao como primeira resposta
            timestamped_content = format_transcription_with_timestamps(transcription_data['segments'])
            project.agent_responses.append(AgentResponse(
                agent_id="transcription",
                agent_name="TRANSCRICAO COM TIMESTAMPS (DETALHADA)",
                content=timestamped_content
            ))

            content_text = transcription_data['text']
            duration_str = format_duration(duration)

            # Etapa 3: Processar com agentes
            self._process_with_agents(
                project, client, config, content_text, timestamped_content,
                duration_str, prefix, base_progress + progress_range * 0.3,
                progress_range * 0.65
            )

            # Limpar audio temporario
            if is_video and audio_path.exists():
                audio_path.unlink()
                try:
                    audio_path.parent.rmdir()
                except Exception:
                    pass

        elif is_document:
            # FLUXO DOCUMENTO: Extrair texto -> Agentes
            self._update_progress(base_progress + progress_range * 0.05,
                                  f"{prefix} - Extraindo conteudo...")

            extracted = extract_file_content(file_path)
            content_text = extracted['text']

            project.transcription_text = content_text

            # Adicionar conteudo extraido como primeira resposta
            project.agent_responses.append(AgentResponse(
                agent_id="extraction",
                agent_name=f"CONTEUDO EXTRAIDO ({extracted['type_label']})",
                content=content_text
            ))

            # Verificar se tem API key para agentes
            if DatabaseManager.has_api_key():
                client = OpenAI(api_key=config.api_key)
                self._process_with_agents(
                    project, client, config, content_text, "", "N/A",
                    prefix, base_progress + progress_range * 0.15,
                    progress_range * 0.8
                )

    def _process_with_agents(self, project: Project, client, config,
                              content_text: str, timestamped_content: str,
                              duration_str: str, prefix: str,
                              agent_base_progress: float, agent_progress_range: float):
        """Processa conteudo com os agentes do pipeline selecionado"""
        # Resolver agentes pelo pipeline
        if self.current_pipeline:
            agents = []
            for aid in self.current_pipeline.agent_ids:
                agent = DatabaseManager.get_agent(aid)
                if agent:
                    agents.append(agent)
        else:
            agents = DatabaseManager.get_agents(ordered=True)
        total_agents = len(agents)
        agent_responses_dict = {}

        for i, agent in enumerate(agents):
            agent_index = i + 1
            progress = agent_base_progress + (agent_progress_range * (i / max(total_agents, 1)))
            self._update_progress(progress, f"{prefix} - {agent.name}")

            # Preparar input base para o agente
            agent_input = f"""TIPO DE ARQUIVO: {project.format}
DURACAO: {duration_str}

CONTEUDO COM TIMESTAMPS:
{timestamped_content if timestamped_content else '(Nao disponivel para este tipo de arquivo)'}

CONTEUDO COMPLETO:
{content_text}"""

            # Substituir variaveis no prompt do agente
            processed_prompt = agent.prompt
            processed_prompt = processed_prompt.replace("{transcricao}", content_text)
            processed_prompt = processed_prompt.replace("{transcricao_timestamps}", timestamped_content or content_text)
            processed_prompt = processed_prompt.replace("{duracao}", duration_str)

            # Substituir variaveis de respostas de agentes anteriores
            for prev_idx, prev_response in agent_responses_dict.items():
                var_name = f"{{response_agente_{prev_idx:02d}}}"
                processed_prompt = processed_prompt.replace(var_name, prev_response)

            # Chamar agente
            response = client.chat.completions.create(
                model=agent.model,
                messages=[
                    {"role": "system", "content": processed_prompt},
                    {"role": "user", "content": agent_input}
                ],
                temperature=0.7,
                max_tokens=4000
            )

            response_content = response.choices[0].message.content
            agent_responses_dict[agent_index] = response_content

            full_prompt_used = f"""=== SYSTEM PROMPT ===
{processed_prompt}

=== USER INPUT ===
{agent_input}"""

            project.agent_responses.append(AgentResponse(
                agent_id=agent.id,
                agent_name=agent.name,
                content=response_content,
                prompt_used=full_prompt_used
            ))

    def _update_progress(self, progress: float, message: str):
        """Atualiza progresso na UI"""
        self.after(0, lambda: self.progress_widget.update_progress(progress, message))

    def _show_success(self):
        """Mostra resultado de sucesso"""
        self.progress_widget.pack_forget()
        self.result_widget.pack()
        msg = "SUCESSO AO PROCESSAR"
        if len(self.current_projects) > 1:
            msg += f" ({len(self.current_projects)} ARQUIVOS)"
        msg += " - CLIQUE ABAIXO PARA VER O RESULTADO"
        self.result_widget.show_success(msg)

    def _show_error(self, message: str):
        """Mostra resultado de erro"""
        self.progress_widget.pack_forget()
        self.result_widget.pack()
        self.result_widget.show_error(message)
        self.file_list_panel.set_enabled(True)

    def _access_result(self):
        """Abre tela de resultado do ultimo projeto processado"""
        if self.last_processed_project:
            self.app.show_agent_responses(self.last_processed_project)

    def _on_project_click(self, project: Project):
        """Callback ao clicar em um projeto"""
        self.app.show_agent_responses(project)

    def refresh(self):
        """Atualiza tela"""
        self.recents_table.refresh()


class HistoryScreen(ctk.CTkFrame):
    """Tela de historico - Tela #06"""

    def __init__(self, parent, app):
        super().__init__(parent, fg_color=COLORS["bg_main"])
        self.app = app

        self._create_widgets()

    def _create_widgets(self):
        """Cria widgets"""
        # Header
        header = HeaderFrame(
            self,
            on_history_click=lambda: None,
            on_settings_click=self.app.show_settings,
            show_history_btn=True
        )
        header.pack(fill="x")

        # Botao voltar
        back_frame = ctk.CTkFrame(self, fg_color="transparent")
        back_frame.pack(fill="x", padx=30, pady=(20, 0))

        ctk.CTkButton(
            back_frame,
            text="< Voltar",
            width=80,
            height=30,
            fg_color="transparent",
            text_color=COLORS["primary"],
            hover_color=COLORS["primary_light"],
            font=ctk.CTkFont(size=13),
            command=self.app.show_home
        ).pack(side="left")

        # Container principal
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=30, pady=20)

        # Titulo
        ctk.CTkLabel(
            main,
            text="Historico de processos",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(pady=(0, 20))

        # Tabela
        self.table_frame = ctk.CTkFrame(main, fg_color=COLORS["bg_card"], corner_radius=8, border_width=1, border_color=COLORS["border"])
        self.table_frame.pack(fill="both", expand=True)

        # Header da tabela
        self._create_table_header()

        # Scrollable frame para rows
        self.scroll_frame = ctk.CTkScrollableFrame(
            self.table_frame,
            fg_color="transparent"
        )
        self.scroll_frame.pack(fill="both", expand=True, padx=2, pady=2)

    def _create_table_header(self):
        """Cria header da tabela"""
        header = ctk.CTkFrame(self.table_frame, fg_color=COLORS["table_header"], height=40)
        header.pack(fill="x", padx=2, pady=(2, 0))
        header.pack_propagate(False)

        columns = [("Titulo", 0.4), ("Tamanho", 0.15), ("Duracao", 0.15), ("Data", 0.3)]

        for col_name, weight in columns:
            ctk.CTkLabel(
                header,
                text=col_name,
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color=COLORS["primary"],
                anchor="w"
            ).pack(side="left", fill="both", expand=True, padx=15, pady=8)

    def refresh(self):
        """Atualiza lista"""
        # Limpar rows
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()

        projects = DatabaseManager.get_projects()

        if not projects:
            ctk.CTkLabel(
                self.scroll_frame,
                text="Nenhum projeto encontrado",
                font=ctk.CTkFont(size=14),
                text_color=COLORS["text_secondary"]
            ).pack(pady=50)
            return

        for project in projects:
            self._create_row(project)

    def _create_row(self, project: Project):
        """Cria linha da tabela"""
        row = ctk.CTkFrame(self.scroll_frame, fg_color="transparent", height=45, cursor="hand2")
        row.pack(fill="x", pady=1)
        row.pack_propagate(False)

        row.bind("<Enter>", lambda e, r=row: r.configure(fg_color=COLORS["table_hover"]))
        row.bind("<Leave>", lambda e, r=row: r.configure(fg_color="transparent"))
        row.bind("<Button-1>", lambda e, p=project: self.app.show_agent_responses(p))

        duration_str = format_duration(project.duration_seconds) if project.duration_seconds else "--:--"
        date_str = project.created_at.strftime("%d/%m/%Y")

        data = [
            (project.file_name[:50] + "..." if len(project.file_name) > 50 else project.file_name, 0.4),
            (f"{project.file_size_mb:.0f} MB", 0.15),
            (duration_str, 0.15),
            (date_str, 0.3)
        ]

        for text, weight in data:
            label = ctk.CTkLabel(
                row,
                text=text,
                font=ctk.CTkFont(size=13),
                text_color=COLORS["text_primary"],
                anchor="w"
            )
            label.pack(side="left", fill="both", expand=True, padx=15, pady=10)
            label.bind("<Button-1>", lambda e, p=project: self.app.show_agent_responses(p))


class AgentResponsesScreen(ctk.CTkFrame):
    """Tela de respostas dos agentes - Tela #05"""

    def __init__(self, parent, app, project: Project):
        super().__init__(parent, fg_color=COLORS["bg_main"])
        self.app = app
        self.project = project
        self.selected_response_idx = 0

        self._create_widgets()

    def _create_widgets(self):
        """Cria widgets"""
        # Header
        header = HeaderFrame(
            self,
            on_history_click=self.app.show_history,
            on_settings_click=self.app.show_settings
        )
        header.pack(fill="x")

        # Botao voltar
        back_frame = ctk.CTkFrame(self, fg_color="transparent")
        back_frame.pack(fill="x", padx=30, pady=(20, 0))

        ctk.CTkButton(
            back_frame,
            text="< Voltar",
            width=80,
            height=30,
            fg_color="transparent",
            text_color=COLORS["primary"],
            hover_color=COLORS["primary_light"],
            font=ctk.CTkFont(size=13),
            command=self.app.show_home
        ).pack(side="left")

        # Titulo
        ctk.CTkLabel(
            back_frame,
            text="Respostas dos agentes",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left", padx=20)

        # Container principal dividido 30/70
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=30, pady=20)

        # Lista de agentes (30%)
        agents_frame = ctk.CTkFrame(main, fg_color=COLORS["bg_card"], corner_radius=8, width=300, border_width=1, border_color=COLORS["border"])
        agents_frame.pack(side="left", fill="y", padx=(0, 15))
        agents_frame.pack_propagate(False)

        self.agents_list = ctk.CTkScrollableFrame(agents_frame, fg_color="transparent")
        self.agents_list.pack(fill="both", expand=True, padx=5, pady=5)

        # Area de conteudo (70%)
        content_frame = ctk.CTkFrame(main, fg_color=COLORS["bg_card"], corner_radius=8, border_width=1, border_color=COLORS["border"])
        content_frame.pack(side="left", fill="both", expand=True)

        # Header da resposta (titulo + botao ver prompt)
        response_header = ctk.CTkFrame(content_frame, fg_color="transparent")
        response_header.pack(fill="x", padx=20, pady=(15, 5))

        # Titulo da resposta
        self.response_title = ctk.CTkLabel(
            response_header,
            text="",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS["primary"],
            anchor="w"
        )
        self.response_title.pack(side="left", fill="x", expand=True)

        # Botao ver prompt usado
        self.view_prompt_btn = ctk.CTkButton(
            response_header,
            text="Ver prompt usado",
            width=140,
            height=32,
            fg_color=COLORS["button_yellow"],
            text_color=COLORS["text_primary"],
            hover_color=COLORS["button_yellow_hover"],
            font=ctk.CTkFont(size=12),
            command=self._show_prompt_modal
        )
        self.view_prompt_btn.pack(side="right")

        # Separador
        ctk.CTkFrame(content_frame, height=1, fg_color=COLORS["border"]).pack(fill="x", padx=20)

        # Texto da resposta
        self.response_text = ctk.CTkTextbox(
            content_frame,
            font=ctk.CTkFont(family="Consolas", size=12),
            text_color=COLORS["text_primary"],
            fg_color="transparent",
            wrap="word"
        )
        self.response_text.pack(fill="both", expand=True, padx=20, pady=15)

        # Preencher lista de agentes
        self._populate_agents_list()

        # Selecionar primeiro
        if self.project.agent_responses:
            self._select_response(0)

    def _populate_agents_list(self):
        """Preenche lista de agentes"""
        for i, response in enumerate(self.project.agent_responses):
            self._create_agent_item(i, response)

    def _create_agent_item(self, idx: int, response: AgentResponse):
        """Cria item de agente na lista"""
        item = ctk.CTkFrame(
            self.agents_list,
            fg_color="transparent",
            corner_radius=8,
            cursor="hand2"
        )
        item.pack(fill="x", pady=3)

        content = ctk.CTkFrame(item, fg_color="transparent")
        content.pack(fill="x", padx=10, pady=10)

        # Icone
        ctk.CTkLabel(
            content,
            text="[AI]",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=COLORS["primary"]
        ).pack(side="left", padx=(0, 10))

        # Nome
        name_label = ctk.CTkLabel(
            content,
            text=response.agent_name,
            font=ctk.CTkFont(size=13),
            text_color=COLORS["text_primary"],
            anchor="w"
        )
        name_label.pack(side="left", fill="x", expand=True)

        # Bind para selecao
        for widget in [item, content, name_label]:
            widget.bind("<Button-1>", lambda e, i=idx: self._select_response(i))
            widget.bind("<Enter>", lambda e, it=item: it.configure(fg_color=COLORS["table_hover"]) if self.selected_response_idx != idx else None)
            widget.bind("<Leave>", lambda e, it=item, i=idx: it.configure(fg_color="transparent") if self.selected_response_idx != i else None)

        # Guardar referencia
        if not hasattr(self, '_agent_items'):
            self._agent_items = []
        self._agent_items.append(item)

    def _select_response(self, idx: int):
        """Seleciona uma resposta"""
        self.selected_response_idx = idx

        # Atualizar visual da lista
        for i, item in enumerate(self._agent_items):
            if i == idx:
                item.configure(fg_color=COLORS["primary_light"])
            else:
                item.configure(fg_color="transparent")

        # Atualizar conteudo
        response = self.project.agent_responses[idx]
        self.response_title.configure(text=response.agent_name)

        self.response_text.configure(state="normal")
        self.response_text.delete("1.0", "end")
        self.response_text.insert("1.0", response.content)
        self.response_text.configure(state="disabled")

        # Mostrar/esconder botao de ver prompt
        # Esconder para transcricao (que nao tem prompt) ou respostas antigas sem prompt
        has_prompt = hasattr(response, 'prompt_used') and response.prompt_used
        if has_prompt:
            self.view_prompt_btn.pack(side="right")
        else:
            self.view_prompt_btn.pack_forget()

    def _show_prompt_modal(self):
        """Abre modal mostrando o prompt usado pelo agente"""
        if self.selected_response_idx >= len(self.project.agent_responses):
            return

        response = self.project.agent_responses[self.selected_response_idx]

        # Verificar se tem prompt
        if not hasattr(response, 'prompt_used') or not response.prompt_used:
            return

        # Criar modal
        PromptViewerModal(self, response.agent_name, response.prompt_used)


class PromptViewerModal(ctk.CTkToplevel):
    """Modal para visualizar o prompt usado pelo agente"""

    def __init__(self, parent, agent_name: str, prompt_content: str):
        super().__init__(parent)
        self.prompt_content = prompt_content

        self.title(f"Prompt usado: {agent_name}")
        self.geometry("800x600")
        self.resizable(True, True)
        self.configure(fg_color=COLORS["bg_main"])

        # Centralizar
        self.update_idletasks()
        x = (self.winfo_screenwidth() - 800) // 2
        y = (self.winfo_screenheight() - 600) // 2
        self.geometry(f"+{x}+{y}")

        self.transient(parent)
        self.grab_set()

        self._create_widgets()

    def _create_widgets(self):
        """Cria widgets"""
        main = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=12)
        main.pack(fill="both", expand=True, padx=20, pady=20)

        # Header
        header = ctk.CTkFrame(main, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=(20, 10))

        ctk.CTkLabel(
            header,
            text="Prompt usado pelo agente (debug)",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")

        # Botao copiar
        ctk.CTkButton(
            header,
            text="Copiar",
            width=80,
            height=32,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=12),
            command=self._copy_to_clipboard
        ).pack(side="right")

        # Texto do prompt (somente leitura)
        self.prompt_text = ctk.CTkTextbox(
            main,
            font=ctk.CTkFont(family="Consolas", size=11),
            text_color=COLORS["text_primary"],
            fg_color=COLORS["bg_main"],
            border_color=COLORS["border"],
            border_width=1,
            wrap="word"
        )
        self.prompt_text.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        self.prompt_text.insert("1.0", self.prompt_content)
        self.prompt_text.configure(state="disabled")

        # Botao fechar
        ctk.CTkButton(
            main,
            text="Fechar",
            width=100,
            height=40,
            fg_color=COLORS["border_dark"],
            text_color=COLORS["text_primary"],
            hover_color=COLORS["border"],
            font=ctk.CTkFont(size=14),
            command=self.destroy
        ).pack(pady=(0, 20))

    def _copy_to_clipboard(self):
        """Copia o prompt para a area de transferencia"""
        self.clipboard_clear()
        self.clipboard_append(self.prompt_content)

        # Feedback visual (mudar texto do botao temporariamente)
        # Nao implementado para simplicidade - o usuario pode usar Ctrl+A, Ctrl+C


class SettingsWindow(ctk.CTkToplevel):
    """Janela de configuracoes - Telas #01.1 e #01.2"""

    def __init__(self, parent, show_api_warning: bool = False):
        super().__init__(parent)
        self.show_api_warning = show_api_warning
        self.current_tab = "llm"

        self.title("Configuracoes do Sistema")
        self.geometry("700x600")
        self.resizable(False, False)
        self.configure(fg_color=COLORS["bg_main"])

        # Centralizar
        self.update_idletasks()
        x = (self.winfo_screenwidth() - 700) // 2
        y = (self.winfo_screenheight() - 600) // 2
        self.geometry(f"+{x}+{y}")

        self.transient(parent)
        self.grab_set()

        self._create_widgets()

    def _create_widgets(self):
        """Cria widgets"""
        # Titulo
        ctk.CTkLabel(
            self,
            text="Configuracoes do Sistema",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(pady=(25, 20))

        # Container principal com abas laterais
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=25, pady=(0, 25))

        # Menu lateral (abas)
        menu = ctk.CTkFrame(main, fg_color="transparent", width=150)
        menu.pack(side="left", fill="y", padx=(0, 20))
        menu.pack_propagate(False)

        self.tab_buttons = {}

        # Aba LLM Config
        self.tab_buttons["llm"] = ctk.CTkButton(
            menu,
            text="LLM Config",
            width=140,
            height=40,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=14),
            command=lambda: self._switch_tab("llm")
        )
        self.tab_buttons["llm"].pack(pady=(0, 5))

        # Aba Agents
        self.tab_buttons["agents"] = ctk.CTkButton(
            menu,
            text="Agents",
            width=140,
            height=40,
            fg_color="transparent",
            text_color=COLORS["text_primary"],
            hover_color=COLORS["primary_light"],
            font=ctk.CTkFont(size=14),
            command=lambda: self._switch_tab("agents")
        )
        self.tab_buttons["agents"].pack(pady=(0, 5))

        # Aba Pipelines
        self.tab_buttons["pipelines"] = ctk.CTkButton(
            menu,
            text="Pipelines",
            width=140,
            height=40,
            fg_color="transparent",
            text_color=COLORS["text_primary"],
            hover_color=COLORS["primary_light"],
            font=ctk.CTkFont(size=14),
            command=lambda: self._switch_tab("pipelines")
        )
        self.tab_buttons["pipelines"].pack()

        # Container de conteudo
        self.content = ctk.CTkFrame(main, fg_color=COLORS["bg_card"], corner_radius=12, border_width=1, border_color=COLORS["border"])
        self.content.pack(side="left", fill="both", expand=True)

        # Frames de conteudo
        self.llm_frame = self._create_llm_frame()
        self.agents_frame = self._create_agents_frame()
        self.pipelines_frame = self._create_pipelines_frame()

        # Mostrar tab inicial
        self._switch_tab("llm")

    def _switch_tab(self, tab: str):
        """Alterna entre abas"""
        self.current_tab = tab

        # Atualizar botoes
        for name, btn in self.tab_buttons.items():
            if name == tab:
                btn.configure(fg_color=COLORS["primary"], text_color=COLORS["text_on_primary"])
            else:
                btn.configure(fg_color="transparent", text_color=COLORS["text_primary"])

        # Mostrar conteudo
        self.llm_frame.pack_forget()
        self.agents_frame.pack_forget()
        self.pipelines_frame.pack_forget()

        if tab == "llm":
            self.llm_frame.pack(fill="both", expand=True, padx=20, pady=20)
        elif tab == "agents":
            self.agents_frame.pack(fill="both", expand=True, padx=20, pady=20)
        elif tab == "pipelines":
            self._refresh_pipelines_list()
            self.pipelines_frame.pack(fill="both", expand=True, padx=20, pady=20)

    def _create_llm_frame(self) -> ctk.CTkFrame:
        """Cria frame de configuracao LLM"""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")

        config = DatabaseManager.get_config()

        # API Key
        ctk.CTkLabel(
            frame,
            text="OpenAI API Key",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))

        self.api_entry = ctk.CTkEntry(
            frame,
            placeholder_text="sk-proj-...",
            show="*",
            height=40,
            border_color=COLORS["error"] if self.show_api_warning and not config.api_key else COLORS["border"],
            fg_color=COLORS["bg_main"]
        )
        self.api_entry.pack(fill="x", pady=(0, 5))
        if config.api_key:
            self.api_entry.insert(0, config.api_key)

        # Toggle mostrar
        self.show_key_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            frame,
            text="Mostrar API Key",
            variable=self.show_key_var,
            command=self._toggle_api_visibility,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 20))

        # Modelo de transcricao
        ctk.CTkLabel(
            frame,
            text="Modelo de transcricao",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))

        self.whisper_combo = ctk.CTkComboBox(
            frame,
            values=WHISPER_MODELS,
            height=40,
            border_color=COLORS["border"],
            fg_color=COLORS["bg_main"],
            button_color=COLORS["primary"],
            button_hover_color=COLORS["primary_hover"]
        )
        self.whisper_combo.set(config.whisper_model)
        self.whisper_combo.pack(fill="x", pady=(0, 30))

        # Botao salvar
        ctk.CTkButton(
            frame,
            text="Salvar",
            width=120,
            height=40,
            fg_color=COLORS["success"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["button_green_hover"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._save_llm_config
        ).pack(side="right")

        return frame

    def _create_agents_frame(self) -> ctk.CTkFrame:
        """Cria frame de agentes"""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")

        # Lista de agentes
        self.agents_list_frame = ctk.CTkScrollableFrame(frame, fg_color="transparent", height=300)
        self.agents_list_frame.pack(fill="both", expand=True, pady=(0, 15))

        self._refresh_agents_list()

        # Botao adicionar
        ctk.CTkButton(
            frame,
            text="ADICIONAR NOVO AGENTE",
            height=45,
            fg_color=COLORS["success"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["button_green_hover"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._add_agent
        ).pack(fill="x")

        return frame

    def _refresh_agents_list(self):
        """Atualiza lista de agentes"""
        for widget in self.agents_list_frame.winfo_children():
            widget.destroy()

        agents = DatabaseManager.get_agents(ordered=True)
        total_agents = len(agents)

        for idx, agent in enumerate(agents):
            self._create_agent_item(agent, idx, total_agents)

    def _create_agent_item(self, agent: Agent, index: int, total: int):
        """Cria item de agente com ordem, nome e botoes"""
        item = ctk.CTkFrame(self.agents_list_frame, fg_color=COLORS["bg_main"], corner_radius=8, border_width=1, border_color=COLORS["border"])
        item.pack(fill="x", pady=5)

        # Linha superior: Ordem + Nome + Botoes de reordenacao
        top_row = ctk.CTkFrame(item, fg_color="transparent")
        top_row.pack(fill="x", padx=15, pady=(12, 5))

        # Numero da ordem (indice + 1)
        order_num = index + 1
        ctk.CTkLabel(
            top_row,
            text=f"#{order_num:02d}",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["primary"],
            width=40
        ).pack(side="left", padx=(0, 5))

        # Botoes de reordenacao (subir/descer)
        order_btn_frame = ctk.CTkFrame(top_row, fg_color="transparent")
        order_btn_frame.pack(side="left", padx=(0, 10))

        # Botao subir (desabilitado se primeiro)
        up_btn = ctk.CTkButton(
            order_btn_frame,
            text="^",
            width=28,
            height=28,
            fg_color=COLORS["primary"] if index > 0 else COLORS["border"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"] if index > 0 else COLORS["border"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=lambda a=agent: self._move_agent_up(a) if index > 0 else None
        )
        up_btn.pack(side="left", padx=(0, 2))
        if index == 0:
            up_btn.configure(state="disabled")

        # Botao descer (desabilitado se ultimo)
        down_btn = ctk.CTkButton(
            order_btn_frame,
            text="v",
            width=28,
            height=28,
            fg_color=COLORS["primary"] if index < total - 1 else COLORS["border"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"] if index < total - 1 else COLORS["border"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=lambda a=agent: self._move_agent_down(a) if index < total - 1 else None
        )
        down_btn.pack(side="left")
        if index == total - 1:
            down_btn.configure(state="disabled")

        # Nome (com quebra de linha automatica)
        ctk.CTkLabel(
            top_row,
            text=agent.name,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"],
            anchor="w",
            wraplength=280
        ).pack(side="left", fill="x", expand=True)

        # Tag da variavel disponivel para agentes posteriores
        var_tag = f"{{response_agente_{order_num:02d}}}"
        ctk.CTkLabel(
            top_row,
            text=var_tag,
            font=ctk.CTkFont(size=10),
            text_color=COLORS["text_secondary"],
            fg_color=COLORS["table_header"],
            corner_radius=4,
            padx=5,
            pady=2
        ).pack(side="right", padx=(5, 0))

        # Linha inferior: Botoes de acao
        btn_frame = ctk.CTkFrame(item, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=(0, 12))

        # Deletar (apenas para agentes nao-sistema)
        if not agent.is_system:
            ctk.CTkButton(
                btn_frame,
                text="Deletar",
                width=80,
                height=32,
                fg_color=COLORS["button_yellow"],
                text_color=COLORS["text_primary"],
                hover_color=COLORS["button_yellow_hover"],
                font=ctk.CTkFont(size=12),
                command=lambda a=agent: self._delete_agent(a)
            ).pack(side="left", padx=(0, 8))

        # Configurar
        ctk.CTkButton(
            btn_frame,
            text="Configurar",
            width=90,
            height=32,
            fg_color=COLORS["button_green"],
            text_color=COLORS["text_primary"],
            hover_color=COLORS["button_green_hover"],
            font=ctk.CTkFont(size=12),
            command=lambda a=agent, idx=index: self._configure_agent(a, idx)
        ).pack(side="left")

    def _move_agent_up(self, agent: Agent):
        """Move agente para cima na ordem"""
        DatabaseManager.move_agent_up(agent.id)
        self._refresh_agents_list()

    def _move_agent_down(self, agent: Agent):
        """Move agente para baixo na ordem"""
        DatabaseManager.move_agent_down(agent.id)
        self._refresh_agents_list()

    def _toggle_api_visibility(self):
        """Alterna visibilidade da API key"""
        self.api_entry.configure(show="" if self.show_key_var.get() else "*")

    def _save_llm_config(self):
        """Salva configuracoes LLM"""
        api_key = self.api_entry.get().strip()

        if not api_key:
            self.api_entry.configure(border_color=COLORS["error"])
            return

        if not api_key.startswith("sk-"):
            self.api_entry.configure(border_color=COLORS["error"])
            return

        config = DatabaseManager.get_config()
        config.api_key = api_key
        config.whisper_model = self.whisper_combo.get()
        DatabaseManager.save_config(config)

        self.api_entry.configure(border_color=COLORS["border"])
        self.destroy()

    def _add_agent(self):
        """Abre janela para adicionar agente"""
        # Novo agente sera o ultimo na ordem
        agents = DatabaseManager.get_agents(ordered=True)
        new_index = len(agents)
        AgentConfigWindow(self, None, self._refresh_agents_list, new_index)

    def _configure_agent(self, agent: Agent, agent_index: int = 0):
        """Abre janela para configurar agente"""
        AgentConfigWindow(self, agent, self._refresh_agents_list, agent_index)

    def _delete_agent(self, agent: Agent):
        """Deleta um agente"""
        if DatabaseManager.delete_agent(agent.id):
            self._refresh_agents_list()

    # --- Pipelines ---

    def _create_pipelines_frame(self) -> ctk.CTkFrame:
        """Cria frame de pipelines"""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")

        # Lista de pipelines
        self.pipelines_list_frame = ctk.CTkScrollableFrame(frame, fg_color="transparent", height=300)
        self.pipelines_list_frame.pack(fill="both", expand=True, pady=(0, 15))

        self._refresh_pipelines_list()

        # Botao adicionar
        ctk.CTkButton(
            frame,
            text="ADICIONAR NOVO PIPELINE",
            height=45,
            fg_color=COLORS["success"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["button_green_hover"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._add_pipeline
        ).pack(fill="x")

        return frame

    def _refresh_pipelines_list(self):
        """Atualiza lista de pipelines"""
        for widget in self.pipelines_list_frame.winfo_children():
            widget.destroy()

        pipelines = DatabaseManager.get_pipelines()

        for pipeline in pipelines:
            self._create_pipeline_item(pipeline)

    def _create_pipeline_item(self, pipeline: Pipeline):
        """Cria item de pipeline na lista"""
        item = ctk.CTkFrame(self.pipelines_list_frame, fg_color=COLORS["bg_main"], corner_radius=8, border_width=1, border_color=COLORS["border"])
        item.pack(fill="x", pady=5)

        # Linha superior: Nome + badge sistema
        top_row = ctk.CTkFrame(item, fg_color="transparent")
        top_row.pack(fill="x", padx=15, pady=(12, 5))

        ctk.CTkLabel(
            top_row,
            text=pipeline.name,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"],
            anchor="w"
        ).pack(side="left", fill="x", expand=True)

        if pipeline.is_system:
            ctk.CTkLabel(
                top_row,
                text=" SISTEMA ",
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color=COLORS["text_on_primary"],
                fg_color=COLORS["primary"],
                corner_radius=4
            ).pack(side="right")

        # Lista de agentes do pipeline
        agents_row = ctk.CTkFrame(item, fg_color="transparent")
        agents_row.pack(fill="x", padx=15, pady=(0, 5))

        agent_names = []
        for aid in pipeline.agent_ids:
            agent = DatabaseManager.get_agent(aid)
            if agent:
                agent_names.append(agent.name)
            else:
                agent_names.append(f"({aid})")

        agents_text = " -> ".join(agent_names) if agent_names else "(nenhum agente)"
        ctk.CTkLabel(
            agents_row,
            text=agents_text,
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"],
            anchor="w",
            wraplength=400
        ).pack(anchor="w")

        # Botoes de acao
        btn_frame = ctk.CTkFrame(item, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=(0, 12))

        if not pipeline.is_system:
            ctk.CTkButton(
                btn_frame,
                text="Deletar",
                width=80,
                height=32,
                fg_color=COLORS["button_yellow"],
                text_color=COLORS["text_primary"],
                hover_color=COLORS["button_yellow_hover"],
                font=ctk.CTkFont(size=12),
                command=lambda p=pipeline: self._delete_pipeline(p)
            ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_frame,
            text="Configurar",
            width=90,
            height=32,
            fg_color=COLORS["button_green"],
            text_color=COLORS["text_primary"],
            hover_color=COLORS["button_green_hover"],
            font=ctk.CTkFont(size=12),
            command=lambda p=pipeline: self._configure_pipeline(p)
        ).pack(side="left")

    def _add_pipeline(self):
        """Abre janela para adicionar pipeline"""
        PipelineConfigWindow(self, None, self._refresh_pipelines_list)

    def _configure_pipeline(self, pipeline: Pipeline):
        """Abre janela para configurar pipeline"""
        PipelineConfigWindow(self, pipeline, self._refresh_pipelines_list)

    def _delete_pipeline(self, pipeline: Pipeline):
        """Deleta um pipeline"""
        if DatabaseManager.delete_pipeline(pipeline.id):
            self._refresh_pipelines_list()


class AgentConfigWindow(ctk.CTkToplevel):
    """Janela de configuracao de agente"""

    def __init__(self, parent, agent: Optional[Agent], on_save_callback, agent_index: int = 0):
        super().__init__(parent)
        self.agent = agent
        self.on_save = on_save_callback
        self.is_new = agent is None
        self.agent_index = agent_index  # Indice do agente na ordem (0-based)

        title = "Adicionar Agente" if self.is_new else f"Prompt do agente: {agent.name}"
        self.title(title)
        self.geometry("700x650")
        self.resizable(False, False)
        self.configure(fg_color=COLORS["bg_main"])

        self.update_idletasks()
        x = (self.winfo_screenwidth() - 700) // 2
        y = (self.winfo_screenheight() - 650) // 2
        self.geometry(f"+{x}+{y}")

        self.transient(parent)
        self.grab_set()

        self._create_widgets()

    def _create_widgets(self):
        """Cria widgets"""
        main = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=12)
        main.pack(fill="both", expand=True, padx=25, pady=25)

        # Container com padding
        container = ctk.CTkFrame(main, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=20)

        # Scrollable frame para o conteudo principal
        scroll_frame = ctk.CTkScrollableFrame(
            container,
            fg_color="transparent",
            scrollbar_button_color=COLORS["border_dark"],
            scrollbar_button_hover_color=COLORS["primary"]
        )
        scroll_frame.pack(fill="both", expand=True, pady=(0, 15))

        # Nome do agente
        ctk.CTkLabel(
            scroll_frame,
            text="Nome do agente",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))

        self.name_entry = ctk.CTkEntry(scroll_frame, height=40, border_color=COLORS["border"], fg_color=COLORS["bg_main"])
        self.name_entry.pack(fill="x", pady=(0, 15))
        if self.agent:
            self.name_entry.insert(0, self.agent.name)
            if self.agent.is_system:
                self.name_entry.configure(state="disabled")

        # Modelo
        ctk.CTkLabel(
            scroll_frame,
            text="Modelo do Agente",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))

        self.model_combo = ctk.CTkComboBox(
            scroll_frame,
            values=GPT_MODELS,
            height=40,
            border_color=COLORS["border"],
            fg_color=COLORS["bg_main"],
            button_color=COLORS["primary"],
            button_hover_color=COLORS["primary_hover"]
        )
        self.model_combo.set(self.agent.model if self.agent else "gpt-4o-mini")
        self.model_combo.pack(fill="x", pady=(0, 15))

        # Variaveis disponiveis
        self._create_variables_info(scroll_frame)

        # Prompt
        ctk.CTkLabel(
            scroll_frame,
            text="Prompt de instrucao",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))

        self.prompt_text = ctk.CTkTextbox(
            scroll_frame,
            height=200,
            border_color=COLORS["border"],
            border_width=1,
            font=ctk.CTkFont(size=12),
            fg_color=COLORS["bg_main"],
            text_color=COLORS["text_primary"]
        )
        self.prompt_text.pack(fill="x", pady=(0, 10))
        if self.agent:
            self.prompt_text.insert("1.0", self.agent.prompt)

        # Botoes (fixos no final, fora do scroll)
        btn_frame = ctk.CTkFrame(container, fg_color="transparent")
        btn_frame.pack(fill="x")

        ctk.CTkButton(
            btn_frame,
            text="Cancelar",
            width=100,
            height=40,
            fg_color=COLORS["border_dark"],
            text_color=COLORS["text_primary"],
            hover_color=COLORS["border"],
            command=self.destroy
        ).pack(side="left")

        ctk.CTkButton(
            btn_frame,
            text="Salvar",
            width=100,
            height=40,
            fg_color=COLORS["success"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["button_green_hover"],
            font=ctk.CTkFont(weight="bold"),
            command=self._save
        ).pack(side="right")

    def _create_variables_info(self, parent):
        """Cria secao com informacoes das variaveis disponiveis"""
        var_frame = ctk.CTkFrame(parent, fg_color=COLORS["table_header"], corner_radius=8)
        var_frame.pack(fill="x", pady=(0, 15))

        inner = ctk.CTkFrame(var_frame, fg_color="transparent")
        inner.pack(fill="x", padx=15, pady=12)

        # Titulo
        ctk.CTkLabel(
            inner,
            text="Variaveis disponiveis no prompt:",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=COLORS["primary"]
        ).pack(anchor="w", pady=(0, 8))

        # Variaveis basicas sempre disponiveis
        basic_vars = [
            ("{transcricao}", "Texto completo da transcricao"),
            ("{transcricao_timestamps}", "Transcricao com timestamps"),
            ("{duracao}", "Duracao do video formatada"),
        ]

        for var, desc in basic_vars:
            var_row = ctk.CTkFrame(inner, fg_color="transparent")
            var_row.pack(fill="x", pady=1)

            ctk.CTkLabel(
                var_row,
                text=var,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=COLORS["primary"],
                width=180,
                anchor="w"
            ).pack(side="left")

            ctk.CTkLabel(
                var_row,
                text=f"- {desc}",
                font=ctk.CTkFont(size=11),
                text_color=COLORS["text_secondary"],
                anchor="w"
            ).pack(side="left", fill="x", expand=True)

        # Variaveis de agentes anteriores (se houver)
        if self.agent_index > 0:
            ctk.CTkFrame(inner, height=1, fg_color=COLORS["border"]).pack(fill="x", pady=8)

            ctk.CTkLabel(
                inner,
                text="Respostas de agentes anteriores:",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=COLORS["primary"]
            ).pack(anchor="w", pady=(0, 5))

            agents = DatabaseManager.get_agents(ordered=True)
            for i in range(self.agent_index):
                if i < len(agents):
                    agent = agents[i]
                    var_name = f"{{response_agente_{i+1:02d}}}"

                    var_row = ctk.CTkFrame(inner, fg_color="transparent")
                    var_row.pack(fill="x", pady=1)

                    ctk.CTkLabel(
                        var_row,
                        text=var_name,
                        font=ctk.CTkFont(size=11, weight="bold"),
                        text_color=COLORS["success"],
                        width=180,
                        anchor="w"
                    ).pack(side="left")

                    # Truncar nome se muito longo
                    agent_name = agent.name[:30] + "..." if len(agent.name) > 30 else agent.name
                    ctk.CTkLabel(
                        var_row,
                        text=f"- Resposta: {agent_name}",
                        font=ctk.CTkFont(size=11),
                        text_color=COLORS["text_secondary"],
                        anchor="w"
                    ).pack(side="left", fill="x", expand=True)
        else:
            ctk.CTkLabel(
                inner,
                text="(Este e o primeiro agente - nao ha respostas anteriores disponiveis)",
                font=ctk.CTkFont(size=10, slant="italic"),
                text_color=COLORS["text_secondary"]
            ).pack(anchor="w", pady=(5, 0))

    def _save(self):
        """Salva agente"""
        name = self.name_entry.get().strip()
        model = self.model_combo.get()
        prompt = self.prompt_text.get("1.0", "end").strip()

        if not name or not prompt:
            return

        if self.is_new:
            agent = Agent(name=name, model=model, prompt=prompt)
            DatabaseManager.add_agent(agent)
        else:
            self.agent.name = name
            self.agent.model = model
            self.agent.prompt = prompt
            DatabaseManager.update_agent(self.agent)

        self.on_save()
        self.destroy()


class PipelineConfigWindow(ctk.CTkToplevel):
    """Janela de configuracao de pipeline - criar/editar"""

    def __init__(self, parent, pipeline: Optional[Pipeline], on_save_callback):
        super().__init__(parent)
        self.pipeline = pipeline
        self.on_save = on_save_callback
        self.is_new = pipeline is None

        # Estado interno: lista ordenada de agent_ids selecionados
        self.selected_agent_ids = list(pipeline.agent_ids) if pipeline else []

        title = "Adicionar Pipeline" if self.is_new else f"Configurar Pipeline: {pipeline.name}"
        self.title(title)
        self.geometry("600x550")
        self.resizable(False, False)
        self.configure(fg_color=COLORS["bg_main"])

        self.update_idletasks()
        x = (self.winfo_screenwidth() - 600) // 2
        y = (self.winfo_screenheight() - 550) // 2
        self.geometry(f"+{x}+{y}")

        self.transient(parent)
        self.grab_set()

        self._create_widgets()

    def _create_widgets(self):
        """Cria widgets"""
        main = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=12)
        main.pack(fill="both", expand=True, padx=20, pady=20)

        container = ctk.CTkFrame(main, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=20)

        # Nome do pipeline
        ctk.CTkLabel(
            container,
            text="Nome do Pipeline",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))

        self.name_entry = ctk.CTkEntry(
            container, height=40,
            border_color=COLORS["border"],
            fg_color=COLORS["bg_main"]
        )
        self.name_entry.pack(fill="x", pady=(0, 15))
        if self.pipeline:
            self.name_entry.insert(0, self.pipeline.name)
            if self.pipeline.is_system:
                self.name_entry.configure(state="disabled")

        # Label da lista de agentes
        ctk.CTkLabel(
            container,
            text="Agentes (marque e ordene com ^ v)",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))

        # Lista scrollavel de agentes
        self.agents_scroll = ctk.CTkScrollableFrame(
            container, fg_color="transparent", height=280,
            scrollbar_button_color=COLORS["border_dark"],
            scrollbar_button_hover_color=COLORS["primary"]
        )
        self.agents_scroll.pack(fill="both", expand=True, pady=(0, 15))

        self._rebuild_agent_list()

        # Botoes Salvar/Cancelar
        btn_frame = ctk.CTkFrame(container, fg_color="transparent")
        btn_frame.pack(fill="x")

        ctk.CTkButton(
            btn_frame,
            text="Cancelar",
            width=100, height=40,
            fg_color=COLORS["border_dark"],
            text_color=COLORS["text_primary"],
            hover_color=COLORS["border"],
            command=self.destroy
        ).pack(side="left")

        ctk.CTkButton(
            btn_frame,
            text="Salvar",
            width=100, height=40,
            fg_color=COLORS["success"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["button_green_hover"],
            font=ctk.CTkFont(weight="bold"),
            command=self._save
        ).pack(side="right")

    def _rebuild_agent_list(self):
        """Reconstroi a lista de agentes com checkboxes e botoes de reordenacao"""
        for widget in self.agents_scroll.winfo_children():
            widget.destroy()

        all_agents = DatabaseManager.get_agents(ordered=True)

        # Separar selecionados (na ordem do pipeline) e nao-selecionados
        selected = []
        for aid in self.selected_agent_ids:
            for a in all_agents:
                if a.id == aid:
                    selected.append(a)
                    break

        unselected = [a for a in all_agents if a.id not in self.selected_agent_ids]

        # Renderizar selecionados primeiro
        for idx, agent in enumerate(selected):
            self._create_agent_row(agent, idx, len(selected), is_selected=True)

        # Separador se houver ambos
        if selected and unselected:
            sep = ctk.CTkFrame(self.agents_scroll, height=1, fg_color=COLORS["border"])
            sep.pack(fill="x", pady=8)

        # Renderizar nao-selecionados
        for agent in unselected:
            self._create_agent_row(agent, -1, 0, is_selected=False)

    def _create_agent_row(self, agent: Agent, index: int, total: int, is_selected: bool):
        """Cria uma linha de agente com checkbox e botoes de ordem"""
        row = ctk.CTkFrame(self.agents_scroll, fg_color="transparent")
        row.pack(fill="x", pady=2)

        # Checkbox
        var = ctk.BooleanVar(value=is_selected)
        cb = ctk.CTkCheckBox(
            row,
            text="",
            variable=var,
            fg_color=COLORS["primary"],
            width=24,
            command=lambda a=agent, v=var: self._toggle_agent(a, v.get())
        )
        cb.pack(side="left", padx=(0, 5))

        # Numero da ordem (se selecionado)
        if is_selected:
            ctk.CTkLabel(
                row,
                text=f"#{index + 1:02d}",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=COLORS["primary"],
                width=35
            ).pack(side="left", padx=(0, 5))

            # Botoes reordenacao
            if index > 0:
                ctk.CTkButton(
                    row, text="^", width=28, height=28,
                    fg_color=COLORS["primary"],
                    text_color=COLORS["text_on_primary"],
                    hover_color=COLORS["primary_hover"],
                    font=ctk.CTkFont(size=12, weight="bold"),
                    command=lambda a=agent: self._move_up(a)
                ).pack(side="left", padx=(0, 2))

            if index < total - 1:
                ctk.CTkButton(
                    row, text="v", width=28, height=28,
                    fg_color=COLORS["primary"],
                    text_color=COLORS["text_on_primary"],
                    hover_color=COLORS["primary_hover"],
                    font=ctk.CTkFont(size=12, weight="bold"),
                    command=lambda a=agent: self._move_down(a)
                ).pack(side="left", padx=(0, 5))

        # Nome do agente
        ctk.CTkLabel(
            row,
            text=agent.name,
            font=ctk.CTkFont(size=13),
            text_color=COLORS["text_primary"] if is_selected else COLORS["text_secondary"],
            anchor="w"
        ).pack(side="left", fill="x", expand=True)

    def _toggle_agent(self, agent: Agent, checked: bool):
        """Adiciona ou remove agente da selecao"""
        if checked and agent.id not in self.selected_agent_ids:
            self.selected_agent_ids.append(agent.id)
        elif not checked and agent.id in self.selected_agent_ids:
            self.selected_agent_ids.remove(agent.id)
        self._rebuild_agent_list()

    def _move_up(self, agent: Agent):
        """Move agente para cima na ordem do pipeline"""
        idx = self.selected_agent_ids.index(agent.id)
        if idx > 0:
            self.selected_agent_ids[idx], self.selected_agent_ids[idx - 1] = \
                self.selected_agent_ids[idx - 1], self.selected_agent_ids[idx]
            self._rebuild_agent_list()

    def _move_down(self, agent: Agent):
        """Move agente para baixo na ordem do pipeline"""
        idx = self.selected_agent_ids.index(agent.id)
        if idx < len(self.selected_agent_ids) - 1:
            self.selected_agent_ids[idx], self.selected_agent_ids[idx + 1] = \
                self.selected_agent_ids[idx + 1], self.selected_agent_ids[idx]
            self._rebuild_agent_list()

    def _save(self):
        """Salva pipeline"""
        name = self.name_entry.get().strip()
        if not name:
            return

        if self.is_new:
            pipeline = Pipeline(
                name=name,
                agent_ids=list(self.selected_agent_ids)
            )
            DatabaseManager.add_pipeline(pipeline)
        else:
            self.pipeline.name = name
            self.pipeline.agent_ids = list(self.selected_agent_ids)
            DatabaseManager.update_pipeline(self.pipeline)

        self.on_save()
        self.destroy()


# ============================================================================
# APLICACAO PRINCIPAL
# ============================================================================

class PytubeCopilotApp(ctk.CTk):
    """Aplicacao principal PyTube Copilot"""

    def __init__(self):
        super().__init__()

        self.title(f"{APP_NAME} v{APP_VERSION}")
        self.geometry("1000x700")
        self.minsize(900, 600)

        # Centralizar
        self.update_idletasks()
        x = (self.winfo_screenwidth() - 1000) // 2
        y = (self.winfo_screenheight() - 700) // 2
        self.geometry(f"+{x}+{y}")

        # Tema claro com azul
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        # Container de telas
        self.screen_container = ctk.CTkFrame(self, fg_color="transparent")
        self.screen_container.pack(fill="both", expand=True)

        # Telas
        self.current_screen = None
        self.screens = {}

        # Mostrar home
        self.show_home()

        # Verificar API key
        self.after(500, self._check_api_key)

    def _clear_screen(self):
        """Limpa tela atual"""
        for widget in self.screen_container.winfo_children():
            widget.destroy()
        self.screens.clear()

    def show_home(self):
        """Mostra tela inicial"""
        self._clear_screen()
        home = HomeScreen(self.screen_container, self)
        home.pack(fill="both", expand=True)
        self.screens["home"] = home
        self.current_screen = "home"

    def show_history(self):
        """Mostra tela de historico"""
        self._clear_screen()
        history = HistoryScreen(self.screen_container, self)
        history.pack(fill="both", expand=True)
        history.refresh()
        self.screens["history"] = history
        self.current_screen = "history"

    def show_settings(self, show_warning: bool = False):
        """Mostra janela de configuracoes"""
        if not DatabaseManager.has_api_key():
            show_warning = True
        SettingsWindow(self, show_warning)

    def show_agent_responses(self, project: Project):
        """Mostra tela de respostas dos agentes"""
        self._clear_screen()
        responses = AgentResponsesScreen(self.screen_container, self, project)
        responses.pack(fill="both", expand=True)
        self.screens["responses"] = responses
        self.current_screen = "responses"

    def _check_api_key(self):
        """Verifica se API key esta configurada"""
        if not DatabaseManager.has_api_key():
            self._show_api_key_warning()

    def _show_api_key_warning(self):
        """Mostra alerta de API key"""
        alert = ctk.CTkToplevel(self)
        alert.title("Atencao - Configuracao Necessaria")
        alert.geometry("450x280")
        alert.resizable(False, False)
        alert.configure(fg_color=COLORS["bg_card"])
        alert.transient(self)
        alert.grab_set()
        alert.protocol("WM_DELETE_WINDOW", lambda: None)

        alert.update_idletasks()
        x = (alert.winfo_screenwidth() - 450) // 2
        y = (alert.winfo_screenheight() - 280) // 2
        alert.geometry(f"+{x}+{y}")

        ctk.CTkLabel(
            alert,
            text="!",
            font=ctk.CTkFont(size=48, weight="bold"),
            text_color=COLORS["warning"]
        ).pack(pady=(30, 10))

        ctk.CTkLabel(
            alert,
            text="API Key da OpenAI Nao Configurada!",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLORS["error"]
        ).pack(pady=(0, 10))

        ctk.CTkLabel(
            alert,
            text="Para utilizar o PyTube Copilot, voce precisa\nconfigurar sua OpenAI API Key.",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["text_primary"],
            justify="center"
        ).pack(pady=(0, 5))

        ctk.CTkLabel(
            alert,
            text="Obtenha sua chave em: platform.openai.com",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_secondary"]
        ).pack(pady=(0, 20))

        def open_settings():
            alert.destroy()
            self.show_settings(show_warning=True)

        ctk.CTkButton(
            alert,
            text="Configurar API Key",
            width=180,
            height=45,
            fg_color=COLORS["primary"],
            text_color=COLORS["text_on_primary"],
            hover_color=COLORS["primary_hover"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=open_settings
        ).pack()


def main():
    """Ponto de entrada"""
    app = PytubeCopilotApp()
    app.mainloop()


if __name__ == "__main__":
    main()
